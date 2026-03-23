"""
领星 ERP 经营数据 MCP Server
部署到 Amazon Bedrock AgentCore，通过 Quick Suite 调用

功能：
1. 利润分析（每日/每周/月度，店铺/MSKU 维度）
2. 差评分析（1-3 星 Review 查询与汇总）
3. 店铺列表查询
4. Excel 报告生成 + S3 上传下载
"""

import os
import sys
import json
import logging
from datetime import datetime, timedelta

# ============================================================
# STEP 1: Patch logging（AgentCore /var/ 只读）
# ============================================================
_original_fh_init = logging.FileHandler.__init__


def _patched_fh_init(self, filename, mode="a", encoding=None, delay=False, errors=None):
    if filename.startswith("/var/"):
        filename = "/tmp/" + os.path.basename(filename)
        print(f"[PATCH] Log redirected to: {filename}", flush=True)
    _original_fh_init(self, filename, mode, encoding, delay, errors)


logging.FileHandler.__init__ = _patched_fh_init

# ============================================================
# STEP 2: 环境变量
# ============================================================
os.environ.setdefault("EXCEL_FILES_PATH", "/tmp/excel_files")
os.makedirs(os.environ["EXCEL_FILES_PATH"], exist_ok=True)

S3_BUCKET = os.environ.get("LINGXING_S3_BUCKET", "")
S3_PREFIX = os.environ.get("LINGXING_S3_PREFIX", "lingxing-reports/")
PRESIGNED_URL_EXPIRY = int(os.environ.get("PRESIGNED_URL_EXPIRY", "3600"))

print("=" * 60, flush=True)
print("领星 ERP MCP Server for AgentCore", flush=True)
print(f"Python: {sys.version}", flush=True)
print(f"S3 Bucket: {S3_BUCKET or 'NOT SET'}", flush=True)
print("=" * 60, flush=True)

# ============================================================
# STEP 3: 创建 MCP Server
# ============================================================
from mcp.server.fastmcp import FastMCP

mcp = FastMCP(
    name="lingxing-erp-mcp",
    host="0.0.0.0",
    port=8000,
    stateless_http=True,
)

# 领星客户端（延迟初始化）
_client = None


def _get_client():
    global _client
    if _client is None:
        from lingxing_client import LingxingClient
        _client = LingxingClient()
    return _client


# ============================================================
# STEP 4: 工具定义
# ============================================================

@mcp.tool()
def get_seller_list() -> str:
    """
    查询所有已授权的亚马逊店铺列表。
    返回店铺 ID(sid)、站点、店铺名、状态等信息。
    其他接口需要用到 sid 参数。
    """
    try:
        result = _get_client().get_seller_list()
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


@mcp.tool()
def get_daily_profit(
    start_date: str,
    end_date: str,
    sids: str = "",
    currency: str = "CNY",
) -> str:
    """
    查询每日利润报表（店铺维度）。

    Args:
        start_date: 开始日期，格式 Y-m-d，如 "2026-03-19"
        end_date: 结束日期，格式 Y-m-d，最长跨度 31 天
        sids: 店铺 ID，多个用逗号分隔，如 "110,120"。留空查全部
        currency: 币种，默认 CNY
    """
    try:
        sid_list = [int(s.strip()) for s in sids.split(",") if s.strip()] if sids else None
        result = _get_client().get_profit_by_seller(
            start_date=start_date,
            end_date=end_date,
            sids=sid_list,
            monthly=False,
            currency=currency,
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


@mcp.tool()
def get_weekly_profit_summary(
    weeks_ago: int = 0,
    sids: str = "",
    currency: str = "CNY",
) -> str:
    """
    查询最近一周的利润汇总（店铺维度）。

    Args:
        weeks_ago: 0=本周，1=上周，2=上上周
        sids: 店铺 ID，多个用逗号分隔。留空查全部
        currency: 币种，默认 CNY
    """
    try:
        today = datetime.now()
        # 计算周一和周日
        start_of_week = today - timedelta(days=today.weekday() + 7 * weeks_ago)
        end_of_week = start_of_week + timedelta(days=6)
        if end_of_week > today:
            end_of_week = today

        start_date = start_of_week.strftime("%Y-%m-%d")
        end_date = end_of_week.strftime("%Y-%m-%d")

        sid_list = [int(s.strip()) for s in sids.split(",") if s.strip()] if sids else None
        result = _get_client().get_profit_by_seller(
            start_date=start_date,
            end_date=end_date,
            sids=sid_list,
            monthly=False,
            currency=currency,
        )

        # 添加查询区间信息
        if isinstance(result, dict):
            result["_query_period"] = {"start": start_date, "end": end_date, "weeks_ago": weeks_ago}

        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


@mcp.tool()
def get_monthly_profit_summary(
    start_month: str,
    end_month: str,
    sids: str = "",
    currency: str = "CNY",
) -> str:
    """
    查询月度利润汇总（店铺维度）。

    Args:
        start_month: 开始月份，格式 Y-m，如 "2026-01"
        end_month: 结束月份，格式 Y-m，如 "2026-03"
        sids: 店铺 ID，多个用逗号分隔。留空查全部
        currency: 币种，默认 CNY
    """
    try:
        sid_list = [int(s.strip()) for s in sids.split(",") if s.strip()] if sids else None
        result = _get_client().get_profit_seller_summary(
            start_date=start_month,
            end_date=end_month,
            sids=sid_list,
            currency=currency,
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


@mcp.tool()
def get_profit_by_msku(
    start_date: str,
    end_date: str,
    sids: str = "",
    currency: str = "CNY",
    offset: int = 0,
    length: int = 500,
) -> str:
    """
    查询利润报表（MSKU 维度），可用于分析单品利润。

    Args:
        start_date: 开始日期 Y-m-d
        end_date: 结束日期 Y-m-d，最长跨度 31 天
        sids: 店铺 ID，多个用逗号分隔
        currency: 币种，默认 CNY
        offset: 分页偏移量
        length: 每页数量，上限 10000
    """
    try:
        sid_list = [int(s.strip()) for s in sids.split(",") if s.strip()] if sids else None
        result = _get_client().get_profit_by_msku(
            start_date=start_date,
            end_date=end_date,
            sids=sid_list,
            offset=offset,
            length=length,
            currency=currency,
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


@mcp.tool()
def get_negative_reviews(
    start_date: str,
    end_date: str,
    star: str = "1,2,3",
    sids: str = "",
    offset: int = 0,
    length: int = 200,
) -> str:
    """
    查询差评 Review 列表（默认 1-3 星）。

    Args:
        start_date: 开始日期 Y-m-d
        end_date: 结束日期 Y-m-d
        star: 星级筛选，多个用逗号分隔，默认 "1,2,3"
        sids: 店铺 ID，多个用逗号分隔。留空查全部
        offset: 分页偏移量
        length: 每页数量，上限 200
    """
    try:
        result = _get_client().get_reviews(
            start_date=start_date,
            end_date=end_date,
            star=star,
            sids=sids,
            offset=offset,
            length=length,
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


@mcp.tool()
def get_today_negative_reviews(sids: str = "") -> str:
    """
    查询今日新增差评（1-3 星），用于每日差评监控。

    Args:
        sids: 店铺 ID，多个用逗号分隔。留空查全部
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        result = _get_client().get_reviews(
            start_date=today,
            end_date=today,
            star="1,2,3",
            sids=sids,
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


# ============================================================
# STEP 5: Excel 报告生成 + S3 上传
# ============================================================

@mcp.tool()
def generate_profit_report_excel(
    start_date: str,
    end_date: str,
    sids: str = "",
    currency: str = "CNY",
    filename: str = "",
) -> str:
    """
    生成利润分析 Excel 报告并上传到 S3，返回下载链接。
    包含店铺维度的销售额、成本、毛利润、广告费等核心指标。

    Args:
        start_date: 开始日期 Y-m-d
        end_date: 结束日期 Y-m-d
        sids: 店铺 ID，多个用逗号分隔
        currency: 币种，默认 CNY
        filename: 自定义文件名（可选）
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import boto3
        from urllib.parse import quote

        # 1. 获取数据
        sid_list = [int(s.strip()) for s in sids.split(",") if s.strip()] if sids else None
        result = _get_client().get_profit_by_seller(
            start_date=start_date,
            end_date=end_date,
            sids=sid_list,
            currency=currency,
        )

        records = []
        if isinstance(result, dict):
            records = result.get("data", {}).get("records", [])
            if not records:
                records = result.get("data", [])

        # 2. 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "利润分析"

        # 表头
        headers = [
            "日期", "店铺", "国家", "币种",
            "销量", "销售额", "FBA销售额", "FBM销售额",
            "广告销售额", "广告费", "促销折扣",
            "退款额", "退款率",
            "FBA发货费", "仓储费", "采购成本", "头程成本",
            "合计成本", "毛利润", "毛利率",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 数据行
        for i, r in enumerate(records, 2):
            row_data = [
                r.get("postedDateLocale", ""),
                r.get("storeName", ""),
                r.get("country", ""),
                r.get("currencyCode", ""),
                r.get("totalSalesQuantity", 0),
                r.get("totalSalesAmount", 0),
                r.get("fbaSaleAmount", 0),
                r.get("fbmSaleAmount", 0),
                r.get("totalAdsSales", 0),
                r.get("totalAdsCost", 0),
                r.get("promotionalRebates", 0),
                r.get("totalSalesRefunds", 0),
                r.get("refundsRate", 0),
                r.get("fbaDeliveryFee", 0),
                r.get("totalStorageFee", 0),
                r.get("cgPriceTotal", 0),
                r.get("cgTransportCostsTotal", 0),
                r.get("totalCost", 0),
                r.get("grossProfit", 0),
                r.get("grossRate", 0),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border

        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        # 3. 保存
        if not filename:
            filename = f"利润分析_{start_date}_{end_date}.xlsx"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        excel_path = os.path.join(os.environ["EXCEL_FILES_PATH"], filename)
        wb.save(excel_path)
        print(f"[EXCEL] Created: {excel_path}", flush=True)

        # 4. 上传 S3
        if not S3_BUCKET:
            return json.dumps({
                "success": True,
                "message": f"报告已生成: {filename}（S3 未配置，无法提供下载链接）",
                "records_count": len(records),
            }, ensure_ascii=False)

        s3 = boto3.client("s3")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        s3_key = f"{S3_PREFIX}{ts}_{filename}"
        encoded = quote(filename, safe="")
        disposition = f"attachment; filename*=UTF-8''{encoded}"

        s3.upload_file(
            excel_path, S3_BUCKET, s3_key,
            ExtraArgs={
                "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "ContentDisposition": disposition,
            },
        )

        url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": S3_BUCKET, "Key": s3_key, "ResponseContentDisposition": disposition},
            ExpiresIn=PRESIGNED_URL_EXPIRY,
        )

        return json.dumps({
            "success": True,
            "download_url": url,
            "filename": filename,
            "records_count": len(records),
            "period": f"{start_date} ~ {end_date}",
            "message": f"利润分析报告已生成，共 {len(records)} 条记录。点击链接下载: {url}",
        }, ensure_ascii=False)

    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


@mcp.tool()
def generate_negative_review_report_excel(
    start_date: str,
    end_date: str,
    sids: str = "",
    filename: str = "",
) -> str:
    """
    生成差评分析 Excel 报告并上传到 S3，返回下载链接。
    包含 1-3 星差评的 ASIN、标题、内容、星级、店铺等信息。

    Args:
        start_date: 开始日期 Y-m-d
        end_date: 结束日期 Y-m-d
        sids: 店铺 ID，多个用逗号分隔
        filename: 自定义文件名（可选）
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import boto3
        from urllib.parse import quote

        # 1. 获取差评数据
        result = _get_client().get_reviews(
            start_date=start_date,
            end_date=end_date,
            star="1,2,3",
            sids=sids,
        )

        reviews = result.get("data", []) if isinstance(result, dict) else []
        total = result.get("total", len(reviews)) if isinstance(result, dict) else 0

        # 2. 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "差评分析"

        headers = [
            "评价日期", "ASIN", "MSKU", "星级", "评价标题", "评价内容",
            "买家", "店铺", "国家", "是否VP", "处理状态", "Review ID", "评价链接",
        ]

        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 星级颜色
        star_colors = {
            1: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            2: PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
            3: PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
        }

        status_map = {0: "待处理", 1: "处理中", 2: "已完成"}

        for i, rv in enumerate(reviews, 2):
            seller_names = rv.get("seller_name", [])
            seller_str = ", ".join(seller_names) if isinstance(seller_names, list) else str(seller_names)
            msku_list = rv.get("seller_sku", [])
            msku_str = ", ".join(msku_list) if isinstance(msku_list, list) else str(msku_list)
            star_val = rv.get("last_star", 0)

            row_data = [
                rv.get("review_date", ""),
                rv.get("asin", ""),
                msku_str,
                star_val,
                rv.get("last_title", ""),
                rv.get("last_content", ""),
                rv.get("author", ""),
                seller_str,
                rv.get("marketplace", ""),
                "是" if rv.get("is_vp") else "否",
                status_map.get(rv.get("status", 0), "未知"),
                rv.get("review_id", ""),
                rv.get("review_url", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                # 星级单元格着色
                if col == 4 and star_val in star_colors:
                    cell.fill = star_colors[star_val]
                    cell.font = Font(color="FFFFFF", bold=True)

        for col in ws.columns:
            max_len = max(len(str(c.value or "")[:50]) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # 3. 保存
        if not filename:
            filename = f"差评分析_{start_date}_{end_date}.xlsx"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        excel_path = os.path.join(os.environ["EXCEL_FILES_PATH"], filename)
        wb.save(excel_path)
        print(f"[EXCEL] Created: {excel_path}", flush=True)

        # 4. 上传 S3
        if not S3_BUCKET:
            return json.dumps({
                "success": True,
                "message": f"差评报告已生成: {filename}（S3 未配置）",
                "total_reviews": total,
            }, ensure_ascii=False)

        s3 = boto3.client("s3")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        s3_key = f"{S3_PREFIX}{ts}_{filename}"
        encoded = quote(filename, safe="")
        disposition = f"attachment; filename*=UTF-8''{encoded}"

        s3.upload_file(
            excel_path, S3_BUCKET, s3_key,
            ExtraArgs={
                "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "ContentDisposition": disposition,
            },
        )

        url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": S3_BUCKET, "Key": s3_key, "ResponseContentDisposition": disposition},
            ExpiresIn=PRESIGNED_URL_EXPIRY,
        )

        return json.dumps({
            "success": True,
            "download_url": url,
            "filename": filename,
            "total_reviews": total,
            "period": f"{start_date} ~ {end_date}",
            "message": f"差评分析报告已生成，共 {total} 条差评。点击链接下载: {url}",
        }, ensure_ascii=False)

    except Exception as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


# ============================================================
# STEP 6: 启动
# ============================================================
print("Tools registered:", flush=True)
print("  - get_seller_list", flush=True)
print("  - get_daily_profit", flush=True)
print("  - get_weekly_profit_summary", flush=True)
print("  - get_monthly_profit_summary", flush=True)
print("  - get_profit_by_msku", flush=True)
print("  - get_negative_reviews", flush=True)
print("  - get_today_negative_reviews", flush=True)
print("  - generate_profit_report_excel", flush=True)
print("  - generate_negative_review_report_excel", flush=True)

if __name__ == "__main__":
    print("Starting MCP server on 0.0.0.0:8000/mcp ...", flush=True)
    mcp.run(transport="streamable-http")
